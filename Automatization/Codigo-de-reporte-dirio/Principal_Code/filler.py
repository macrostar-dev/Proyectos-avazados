from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime


def generar_reporte_ko_misma_carpeta(
    archivo_origen: str,
    archivo_destino: str,
    hoja_destino: str,
    hoja_origen: str = "Sheet1",
):
    """
    archivo_origen y archivo_destino se consideran relativos a la carpeta Uadmin:
      - Ej: archivo_origen="Reportes/Grecia.xlsx"
      - Ej: archivo_destino="x.xlsx"
    """

    # Carpeta base = padre de Principal_Code → Uadmin
    BASE_DIR = Path(__file__).resolve().parent.parent

    ruta_origen = BASE_DIR / archivo_origen
    ruta_destino = BASE_DIR / archivo_destino

    print("Archivo origen:", ruta_origen)
    print("Archivo destino:", ruta_destino)

    if not ruta_origen.exists():
        raise FileNotFoundError(f"El archivo de origen NO existe: {ruta_origen}")

    if not ruta_destino.exists():
        raise FileNotFoundError(f"El archivo de destino NO existe: {ruta_destino}")

    wb_origen = load_workbook(ruta_origen, data_only=True)
    wb_destino = load_workbook(ruta_destino)

    ho = wb_origen[hoja_origen]
    hd = wb_destino[hoja_destino]

    columna_filtro = "T"

    mapeo = {
        "I": "A",
        "A": "F",
        "Y": "G",
    }

    columna_base_destino = "A"

    ahora = datetime.now()
    fecha_hoy = ahora.strftime("%d/%m/%Y")
    hora_actual = "3:00:00 pm"

    # Encontrar primera fila libre en destino
    fila_destino = 2
    for fila in range(hd.max_row, 1, -1):
        valor = hd[f"{columna_base_destino}{fila}"].value
        if valor not in (None, ""):
            fila_destino = fila + 1
            break

    fila_inicio_insercion = fila_destino
    print("Primera fila libre en destino detectada:", fila_destino)

    filas_con_ko = 0

    for fila in range(2, ho.max_row + 1):
        valor_filtro = ho[f"{columna_filtro}{fila}"].value
        texto = "" if valor_filtro is None else str(valor_filtro).strip().upper()

        if texto == "KO":
            filas_con_ko += 1

            # Copiar columnas origen -> destino solo si están vacías
            for col_origen, col_destino in mapeo.items():
                valor = ho[f"{col_origen}{fila}"].value
                celda_dest = hd[f"{col_destino}{fila_destino}"]
                if celda_dest.value in (None, ""):
                    celda_dest.value = valor

            # B = fecha
            celda_B = hd[f"B{fila_destino}"]
            if celda_B.value in (None, ""):
                celda_B.value = fecha_hoy

            # C = hora con segundos
            celda_C = hd[f"C{fila_destino}"]
            if celda_C.value in (None, ""):
                celda_C.value = hora_actual

            # E = "Desconectado"
            celda_E = hd[f"E{fila_destino}"]
            if celda_E.value in (None, ""):
                celda_E.value = "Desconectado"

            # H = "más de una hora "
            celda_H = hd[f"H{fila_destino}"]
            if celda_H.value in (None, ""):
                celda_H.value = "más de una hora "

            fila_destino += 1

    # Poner total en D
    if filas_con_ko > 0:
        for fila in range(fila_inicio_insercion, fila_inicio_insercion + filas_con_ko):
            hd[f"D{fila}"].value = filas_con_ko

        fila_fin = fila_inicio_insercion + filas_con_ko - 1
        print(
            f"Se insertaron {filas_con_ko} filas nuevas "
            f"a partir de la fila {fila_inicio_insercion} hasta la fila {fila_fin}."
        )
    else:
        print("No se insertaron filas nuevas (no se encontraron KO).")

    wb_destino.save(ruta_destino)
    print("✔ Proceso terminado.")
    print("Archivo modificado:", ruta_destino)
    print("Total filas con KO insertadas:", filas_con_ko)

    return filas_con_ko
