import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Carpeta base = padre de Principal_Code → Uadmin
BASE_DIR = Path(__file__).resolve().parent.parent
CARPETA_TRABAJO_1 = BASE_DIR                     # Uadmin
CARPETA_TRABAJO_2 = BASE_DIR / "Registro_Anual"  # Destino
DOCUMENTO_TRABAJO_1 = BASE_DIR / "Monitoreo Diario U-Admin Grecia-Desamparados.xlsx"

def copiar_archivo_con_fecha_anual(palabra_buscar: str):
    """
    Busca un archivo en Uadmin (BASE_DIR) que contenga 'palabra_buscar'
    en el nombre y crea una COPIA en la carpeta Registro_Anual
    agregando la fecha al nombre.
    """

    palabra_buscar = palabra_buscar.lower().strip()

    if not CARPETA_TRABAJO_1.exists():
        raise FileNotFoundError(f"❌ La carpeta base no existe: {CARPETA_TRABAJO_1}")

    # Crear carpeta destino si no existe
    CARPETA_TRABAJO_2.mkdir(parents=True, exist_ok=True)

    # 1. Buscar archivo en Uadmin
    archivo_encontrado = next(
        (
            f for f in CARPETA_TRABAJO_1.iterdir()
            if f.is_file() and palabra_buscar in f.name.lower()
        ),
        None
    )

    if not archivo_encontrado:
        print(f"❌ No se encontró ningún archivo que contenga: '{palabra_buscar}' en {CARPETA_TRABAJO_1}")
        return None

    print(f"📄 Archivo encontrado: {archivo_encontrado.name}")

    # 2. Fecha
    fecha_hoy = datetime.now().strftime("%Y")
    fecha_para_nombre = fecha_hoy.replace("/", "-")

    # 3. Nuevo nombre
    nuevo_nombre = f"{archivo_encontrado.stem} {fecha_para_nombre}{archivo_encontrado.suffix}"
    ruta_copia = CARPETA_TRABAJO_2 / nuevo_nombre

    # 4. Copiar archivo
    shutil.copy2(archivo_encontrado, ruta_copia)

    print(f"📑 Copia creada en: {ruta_copia}")
    print(f"🗓 Fecha usada (lógica): {fecha_hoy}")

    return ruta_copia

# Ejecución
# print(copiar_archivo_con_fecha(palabra_buscar="Diario"))
from openpyxl import load_workbook

def limpiar_excel_sin_encabezados(ruta_excel):
    wb = load_workbook(ruta_excel)

    for hoja in wb.worksheets:
        if hoja.max_row >= 3:
            hoja.delete_rows(3, hoja.max_row)

    wb.save(ruta_excel)


#limpiar_excel_sin_encabezados(DOCUMENTO_TRABAJO_1)

def actualizacion_anual():
    if datetime.now().strftime("%d/%m") == "1/1":
        copiar_archivo_con_fecha_anual(palabra_buscar="Diario")
        limpiar_excel_sin_encabezados(ruta_excel=DOCUMENTO_TRABAJO_1)
        print("Reporte reseteado")
    else:
        print()

actualizacion_anual()